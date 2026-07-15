#!/usr/bin/env python3
import argparse
import os
import sys
import requests
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

LAT = 45.073475
LON = 7.543461
ENSEMBLE_URL = "https://ensemble-api.open-meteo.com/v1/ensemble"

# Mappatura identica a ens_rivoli.py per coerenza
MODELLI = {
    "ch1": {
        "id_api": "meteoswiss_icon_ch1_ensemble",
        "nome": "ICON-CH1-EPS",
        "orizzonte_ore": 33,
        "is_ensemble": True,
        "giorni": 2
    },
    "ch2": {
        "id_api": "meteoswiss_icon_ch2_ensemble",
        "nome": "ICON-CH2-EPS",
        "orizzonte_ore": 120,
        "is_ensemble": True,
        "giorni": 5
    },
    "d2": {
        "id_api": "icon_d2",
        "nome": "ICON-D2-EPS",
        "orizzonte_ore": 48,
        "is_ensemble": True,
        "giorni": 2
    },
    "arome": {
        "is_ensemble": False
    },
    "icon2i": {
        "is_ensemble": False
    }
}

def fetch_data(modello):
    """Scarica i dati ensemble da Open-Meteo per il modello specificato."""
    if not MODELLI[modello]["is_ensemble"]:
        print(f"ℹ️ {modello.upper()} non è un ensemble. Nessun dato da scaricare.")
        return None

    variabili = "temperature_2m,precipitation,dew_point_2m,relative_humidity_2m,freezing_level_height"
    params = {
        "latitude": LAT,
        "longitude": LON,
        "models": MODELLI[modello]["id_api"],
        "hourly": variabili,
        "forecast_days": MODELLI[modello]["giorni"],
        "timezone": "Europe/Rome"
    }
    
    nome_modello = MODELLI[modello]["nome"]
    print(f"⏳ Download dati da Open-Meteo per {nome_modello} in corso...")
    
    # Retry per robustezza
    for tentativo in range(3):
        try:
            resp = requests.get(ENSEMBLE_URL, params=params, timeout=90)
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            print(f"Tentativo {tentativo+1} fallito per {nome_modello}: {e}")
            if tentativo == 2:
                raise e
            import time; time.sleep(5)

def genera_excel(data, out_path):
    """Elabora i JSON aggregando min/max/media e calcolando prob. pioggia."""
    print(f"📊 Generazione del file Excel: {out_path}")
    df = pd.DataFrame(data['hourly'])
    df['time'] = pd.to_datetime(df['time'])

    summary_df = pd.DataFrame({'Data_Ora': df['time'].dt.strftime('%Y-%m-%d %H:%M')})

    metriche_base = {
        'temperature_2m': ('Temp', '°C'),
        'dew_point_2m': ('DewPoint', '°C'),
        'relative_humidity_2m': ('Umidita', '%'),
        'freezing_level_height': ('ZeroTermico', 'm'),
        'precipitation': ('Precip', 'mm')
    }

    # Calcolo metriche
    for var, (nome, unita) in metriche_base.items():
        cols = [c for c in df.columns if var in c]
        if cols:
            summary_df[f'{nome} Min ({unita})'] = df[cols].min(axis=1)
            summary_df[f'{nome} Media ({unita})'] = df[cols].mean(axis=1)
            summary_df[f'{nome} Max ({unita})'] = df[cols].max(axis=1)

    # Calcolo Probabilità di Precipitazione
    precip_cols = [c for c in df.columns if 'precipitation' in c]
    if precip_cols:
        for p in range(1, 31):
            summary_df[f'Prob Pioggia >{p}mm'] = (df[precip_cols] >= p).sum(axis=1) / len(precip_cols)

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Analisi_Ensemble', index=False)

    # Formattazione
    wb = openpyxl.load_workbook(out_path)
    ws = wb['Analisi_Ensemble']
    ws.freeze_panes = "B2"
    
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alt_row_fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    border = Border(left=Side(style='thin', color='D5D8DC'), right=Side(style='thin', color='D5D8DC'), 
                    top=Side(style='thin', color='D5D8DC'), bottom=Side(style='thin', color='D5D8DC'))
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        intestazione = str(ws.cell(row=1, column=col[0].column).value)
        
        for i, cell in enumerate(col):
            if cell.value: 
                max_length = max(max_length, len(str(cell.value)))
            if i > 0:
                cell.border = border
                cell.alignment = center_align
                if (i + 1) % 2 == 0: 
                    cell.fill = alt_row_fill
                if isinstance(cell.value, (int, float)):
                    if "Prob" in intestazione:
                        cell.number_format = '0.0%'
                    elif "ZeroTermico" in intestazione:
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '0.00'

        ws.column_dimensions[column].width = min(max(len(intestazione) + 2, max_length + 2), 20)

    wb.save(out_path)

def invia_documento_telegram(percorso_file, didascalia):
    """Invia l'Excel tramite le API di Telegram."""
    token = os.getenv("TELEGRAM_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")
    
    if not token or not chat_id:
        print("⚠️ Credenziali Telegram mancanti. Salto l'invio.", file=sys.stderr)
        return

    url = f"https://api.telegram.org/bot{token}/sendDocument"
    print(f"📤 Invio {percorso_file} su Telegram...")
    
    try:
        with open(percorso_file, "rb") as documento:
            resp = requests.post(
                url, 
                data={"chat_id": chat_id, "caption": didascalia}, 
                files={"document": documento},
                timeout=60
            )
            resp.raise_for_status()
        print(f"✅ Inviato su Telegram!")
    except Exception as e:
        print(f"❌ Errore invio Telegram: {e}", file=sys.stderr)

def main():
    parser = argparse.ArgumentParser(description="Generatore Dati Excel Ensemble")
    parser.add_argument("--modello", type=str, choices=["ch1", "ch2", "d2", "arome", "icon2i"], required=True)
    args = parser.parse_args()

    modello_key = args.modello
    
    if non MODELLI[modello_key]["is_ensemble"]:
        print(f"Salto la generazione Excel per {modello_key.upper()} (non è un ensemble)")
        sys.exit(0)
        
    nome_modello = MODELLI[modello_key]["nome"]
    out_path = f"Dati_Ensemble_Rivoli_{modello_key.upper()}.xlsx"

    try:
        dati = fetch_data(modello_key)
        if dati:
            genera_excel(dati, out_path)
            ora = datetime.now().strftime("%d/%m/%Y alle %H:%M")
            didascalia = f"📉 Dati Tabellari {nome_modello}\n📍 Rivoli (TO) - Aggiornato il {ora}"
            invia_documento_telegram(out_path, didascalia)
    except Exception as e:
        print(f"❌ Errore critico per {modello_key}: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
